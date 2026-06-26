"""
碳排放因子搜索数据库
从三个权威数据源加载排放因子，提供统一的关键词搜索接口。

数据来源：
  China_LCA  — 中国产品全生命周期温室气体排放系数集（CCG, 2022）
               6 类：工业产品、能源产品、生活产品、废弃物处理、交通服务、碳汇
  IPCC       — IPCC碳排放因子数据库
               燃料燃烧 CO2 排放系数 + GWP 值 + 外购电力/蒸汽
  UK_DEFRA   — 英国环境部 DEFRA 常用版 2021
               Scope1 燃料、乘用车、Scope2 电力、Scope3 航空/陆路/货运/废物/材料
"""

import json
import logging
import os
import re
import warnings
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

_DB_DIR = Path(os.path.expanduser("~/Desktop/碳排放因子数据库"))
_CHINA_LCA_PATH = _DB_DIR / "中国产品全生命周期温室气体排放系数库.xlsx"
_IPCC_PATH      = _DB_DIR / "IPCC碳排放因子数据库.xls"
_UK_DEFRA_PATH  = _DB_DIR / "英国环境部 EFDB-常用版 2021.xlsm"

_RECORDS: list[dict] = []
_LOADED = False


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r'\s+', ' ', str(v)).strip()


def _num(v) -> Optional[float]:
    try:
        return float(v) if v is not None and str(v).strip() not in ("", "-", "－") else None
    except (ValueError, TypeError):
        return None


def _norm_unit(u: str) -> str:
    u = re.sub(r'[（(][^）)]*[）)]', '', u).strip()
    return re.sub(r'\s+', ' ', u)


# ── China LCA ────────────────────────────────────────────────────

def _load_china_lca() -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping China LCA")
        return []
    if not _CHINA_LCA_PATH.exists():
        log.warning(f"China LCA not found: {_CHINA_LCA_PATH}")
        return []

    recs = []
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(str(_CHINA_LCA_PATH), read_only=True, data_only=True)

        data_sheets = ["工业产品", "能源产品", "生活产品", "废弃物处理", "交通服务", "碳汇"]
        for sname in data_sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            cat1 = cat2 = ""
            for ri, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
                c1, c2, c3 = _clean(row[0]), _clean(row[1]), _clean(row[2]) if len(row) > 2 else ""
                if c1:
                    cat1 = c1
                if c2:
                    cat2 = c2
                if not c3:
                    continue
                up_val = _num(row[3]) if len(row) > 3 else None
                up_unit = _norm_unit(_clean(row[4])) if len(row) > 4 and row[4] else None
                dn_val = _num(row[5]) if len(row) > 5 else None
                dn_unit = _norm_unit(_clean(row[6])) if len(row) > 6 and row[6] else None
                desc = _clean(row[7]) if len(row) > 7 and row[7] else None

                recs.append({
                    "id": f"cn_lca_{sname}_{ri}",
                    "source": "China_LCA",
                    "category": cat1,
                    "subcategory": cat2,
                    "name": c3,
                    "factor": up_val,
                    "unit": up_unit,
                    "factor2": dn_val,
                    "unit2": dn_unit,
                    "description": desc,
                    "year": "2022",
                })
    except Exception as e:
        log.error(f"China LCA load error: {e}")
    log.info(f"China_LCA: {len(recs)} records")
    return recs


# ── IPCC ─────────────────────────────────────────────────────────

def _load_ipcc() -> list[dict]:
    try:
        import xlrd
    except ImportError:
        log.warning("xlrd not installed — skipping IPCC")
        return []
    if not _IPCC_PATH.exists():
        log.warning(f"IPCC file not found: {_IPCC_PATH}")
        return []

    recs = []
    try:
        wb = xlrd.open_workbook(str(_IPCC_PATH), on_demand=True)

        # Sheet 0: CO2 from fuel combustion
        sh = wb.sheet_by_index(0)
        gas_type = src_type = fuel_cat = ""
        for i in range(6, sh.nrows):
            row = sh.row_values(i)
            if row[0]: gas_type = _clean(row[0])
            if row[1]: src_type = _clean(row[1])
            if row[2]: fuel_cat = _clean(row[2])
            fuel = _clean(row[3])
            co2_ef = _num(row[7])
            unit = _clean(row[8]) if len(row) > 8 and row[8] else "kgCO2/TJ"
            if not fuel or co2_ef is None:
                continue
            recs.append({
                "id": f"ipcc_co2_{i}",
                "source": "IPCC",
                "category": f"燃料燃烧-{src_type}",
                "subcategory": fuel_cat,
                "name": fuel,
                "factor": co2_ef,
                "unit": unit,
                "factor2": None, "unit2": None,
                "description": f"{gas_type}排放，{src_type}燃烧",
                "year": "2006",
            })

        # Sheet 3: GWP values
        sh4 = wb.sheet_by_index(3)
        for i in range(2, sh4.nrows):
            row = sh4.row_values(i)
            gas = _clean(row[0])
            gwp = _num(row[1])
            src = _clean(row[2]) if len(row) > 2 and row[2] else "IPCC"
            if not gas or gwp is None:
                continue
            recs.append({
                "id": f"ipcc_gwp_{i}",
                "source": "IPCC",
                "category": "温室气体GWP值",
                "subcategory": "全球增温潜势",
                "name": gas,
                "factor": gwp,
                "unit": "kgCO2e/kg",
                "factor2": None, "unit2": None,
                "description": f"来源：{src}",
                "year": "2001",
            })

        # Sheet 4: 外购电力/蒸汽
        sh5 = wb.sheet_by_index(4)
        label = ""
        for i in range(3, sh5.nrows):
            row = sh5.row_values(i)
            if row[0]: label = _clean(row[0])
            name = _clean(row[1]) if len(row) > 1 and row[1] else ""
            co2e = _num(row[5]) if len(row) > 5 else None
            unit = _clean(row[6]) if len(row) > 6 and row[6] else ""
            if not name or co2e is None:
                continue
            recs.append({
                "id": f"ipcc_grid_{i}",
                "source": "IPCC",
                "category": "外购电力与蒸汽",
                "subcategory": label or "电力/蒸汽",
                "name": name,
                "factor": co2e,
                "unit": unit,
                "factor2": None, "unit2": None,
                "description": None,
                "year": "2006",
            })

    except Exception as e:
        log.error(f"IPCC load error: {e}")
    log.info(f"IPCC: {len(recs)} records")
    return recs


# ── UK DEFRA ──────────────────────────────────────────────────────

_UK_SHEETS = {
    "Fuels":                    "Scope1-燃料",
    "Passenger vehicles":       "Scope1-乘用车",
    "UK electricity":           "Scope2-英国电力",
    "Business travel- air":     "Scope3-商务航空",
    "Business travel- land":    "Scope3-商务陆路",
    "Freighting goods":         "Scope3-货物运输",
    "Waste disposal":           "Scope3-废弃物处置",
    "Material use":             "Scope3-材料",
}


def _load_uk_defra() -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping UK DEFRA")
        return []
    if not _UK_DEFRA_PATH.exists():
        log.warning(f"UK DEFRA not found: {_UK_DEFRA_PATH}")
        return []

    recs = []
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(str(_UK_DEFRA_PATH), read_only=True, data_only=True)

        for sname, cat_label in _UK_SHEETS.items():
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))

            # Find header: row with "kg CO2e" and "Unit"
            header_row = None
            for idx, row in enumerate(rows):
                txt = " ".join(str(v) for v in row if v)
                if "kg CO2e" in txt and "Unit" in txt:
                    header_row = idx
                    break
            if header_row is None:
                continue

            activity = fuel_name = ""
            for row in rows[header_row + 1:]:
                if all(v is None for v in row):
                    continue
                act  = _clean(row[1]) if len(row) > 1 and row[1] else ""
                name = _clean(row[2]) if len(row) > 2 and row[2] else ""
                unit = _clean(row[3]) if len(row) > 3 and row[3] else ""
                co2e = _num(row[4]) if len(row) > 4 else None
                if act:
                    activity = act
                if name:
                    fuel_name = name
                if not unit or co2e is None:
                    continue
                rid = re.sub(r'[^A-Za-z0-9_]', '_', f"uk_{sname}_{fuel_name}_{unit}")
                recs.append({
                    "id": rid,
                    "source": "UK_DEFRA",
                    "category": cat_label,
                    "subcategory": activity,
                    "name": f"{fuel_name} ({unit})",
                    "factor": co2e,
                    "unit": f"kgCO2e/{unit}",
                    "factor2": None, "unit2": None,
                    "description": f"UK DEFRA 2021",
                    "year": "2021",
                })
    except Exception as e:
        log.error(f"UK DEFRA load error: {e}")
    log.info(f"UK_DEFRA: {len(recs)} records")
    return recs


# ── Public API ───────────────────────────────────────────────────

def load_all() -> list[dict]:
    global _RECORDS, _LOADED
    if not _LOADED:
        _RECORDS = _load_china_lca() + _load_ipcc() + _load_uk_defra()
        _LOADED = True
        log.info(f"EF database total: {len(_RECORDS)} records")
    return _RECORDS


def search(
    keyword: str = None,
    category: str = None,
    source: str = None,
    limit: int = 15,
) -> list[dict]:
    """
    搜索排放因子。
      keyword  — 关键词，匹配 name / subcategory / category / description
      category — 一级分类过滤（如 "工业产品", "燃料燃烧", "交通服务"）
      source   — 数据源：China_LCA / IPCC / UK_DEFRA
      limit    — 最多返回条数（默认 15）
    """
    data = load_all()
    res = data

    if source:
        su = source.upper()
        res = [r for r in res if su in r["source"].upper()]

    if category:
        cl = category.lower()
        res = [r for r in res if cl in r["category"].lower() or cl in r["subcategory"].lower()]

    if keyword:
        kw = keyword.lower()
        def _hit(r):
            return (kw in r["name"].lower()
                    or kw in r["subcategory"].lower()
                    or kw in r["category"].lower()
                    or bool(r["description"] and kw in r["description"].lower()))
        res = [r for r in res if _hit(r)]

    res = [r for r in res if r["factor"] is not None]
    return res[:limit]


def execute_search_emission_factors(
    keyword: str = None,
    category: str = None,
    source: str = None,
) -> str:
    """返回 JSON 字符串供 agent 工具调用使用"""
    hits = search(keyword=keyword, category=category, source=source, limit=15)
    if not hits:
        return json.dumps(
            {"message": "未找到匹配的排放因子，请尝试不同关键词或去掉过滤条件"},
            ensure_ascii=False
        )
    out = []
    for r in hits:
        item = {
            "source": r["source"],
            "category": r["category"],
            "name": r["name"],
            "factor": r["factor"],
            "unit": r["unit"],
            "year": r["year"],
        }
        if r["factor2"] is not None:
            item["downstream_factor"] = r["factor2"]
            item["downstream_unit"] = r["unit2"]
        if r["description"]:
            item["note"] = r["description"][:120]
        out.append(item)
    return json.dumps({
        "total_found": len(hits),
        "results": out,
        "tip": "factor=上游/生产排放（China_LCA也提供downstream_factor=使用阶段排放）",
    }, ensure_ascii=False, indent=2)
